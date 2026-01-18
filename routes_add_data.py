# routes_add_data.py
# -------------------------------------
# All /add_data routes extracted cleanly
# -------------------------------------

import os, random, subprocess
from flask import Blueprint, render_template, request, redirect, url_for
from flask_login import login_required
from openpyxl import load_workbook
from markupsafe import Markup
from werkzeug.utils import secure_filename

# import shared functions/objects from your app
from app import app, extract_name_entities, get_field_list, driver
from werkzeug.security import generate_password_hash


bp_add_data = Blueprint("add_data_routes", __name__)


# -----------------------
# /add_data
# -----------------------
@bp_add_data.route('/add_data', methods=['GET'])
@login_required
def add_data():
    fields = get_field_list()
    return render_template("objekt_form_buttons.html",
                           fields=fields, data={}, colors={}, name_entities={})


# -----------------------
# /add_data/upload Excel
# -----------------------
@bp_add_data.route('/add_data/upload', methods=['POST'])
@login_required
def add_data_upload():
    file = request.files.get('file')
    if not file or not file.filename:
        return "❌ No file selected", 400

    path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
    file.save(path)

    wb = load_workbook(path)
    ws = wb.active

    data, colors, name_entities = {}, {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        key = str(row[0]).strip()
        val = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        data[key] = val
        colors[key] = "yellow" if val else "red"
        name_entities[key] = extract_name_entities(val)

    return render_template("objekt_form_buttons.html",
                           fields=list(data.keys()), data=data,
                           colors=colors, name_entities=name_entities)


# -----------------------
# /add_data/extract_entities
# -----------------------
@bp_add_data.route('/add_data/extract_entities', methods=['POST'])
@login_required
def add_data_extract_entities():
    fields = get_field_list()
    data, colors, name_entities = {}, {}, {}

    for key in fields:
        val = request.form.get(key, "").strip()
        ents = extract_name_entities(val)

        data[key] = val
        name_entities[key] = ents
        colors[key] = "green" if ents else "yellow" if val else "red"

    return render_template("objekt_form_buttons.html",
                           fields=fields, data=data,
                           colors=colors, name_entities=name_entities)


# -----------------------
# /add_data/submit
# -----------------------
@bp_add_data.route('/add_data/submit', methods=['POST'])
@login_required
def add_data_submit():
    posted = request.form
    rows = []

    for key, value in posted.items():
        if key.startswith("custom_property_"):
            i = key.split("_")[-1]
            name = value.strip()
            val  = posted.get(f"custom_value_{i}", "").strip() or "No data"
            status = posted.get(f"traffic_custom_{i}", "red")
            rows.append([name or "No data", val, status, ""])
        elif key.startswith("traffic_") or key in ("source_node_id","relation_name","category","nodeType"):
            continue
        else:
            val = value.strip() or "No data"
            status = posted.get(f"traffic_{key}", "red")
            rows.append([key, val, status, ""])

    node_type = posted.get("nodeType", posted.get("category", "article"))
    rows.append(["nodeType", node_type, "", ""])

    # assemble into 1 string for Java input
    out = "Property=text=Status=Name Entity\n"
    for r in rows:
        out += "=".join(str(x) for x in r) + "\n"

    # generate object id
    objekt_id = f"{posted.get('category','article')}_{random.randint(1,100)}"

    CSV_DIR = "/app/marburg-project/dataset/german/input/"
    JAR = "/app/marburg-project/target/QuestionGrammarGenerator.jar"

    cmd = [
        "java","-jar",JAR,
        "CREATE_FROM_STRING",
        CSV_DIR,
        os.getenv("NEO4J_URI","bolt://neo4j:7687"),
        os.getenv("NEO4J_USER","neo4j"),
        os.getenv("NEO4J_PASSWORD","password"),
        out,
        "FALSE"
    ]

    try:
        subprocess.run(cmd, capture_output=True, text=True, check=True)
    except subprocess.CalledProcessError as e:
        return f"<h3>ERROR: {e.stderr}</h3>"

    return redirect(url_for("add_data_routes.add_data"))


# -----------------------
# /add_data/open_entity_form
# -----------------------
@bp_add_data.route('/add_data/open_entity_form')
@login_required
def add_data_open_entity_form():
    relation = request.args.get("relation")
    source = request.args.get("source_node_id")
    entity  = request.args.get("entity_name")

    fields = get_field_list()
    data, colors, names = {}, {}, {}

    field = f"Verknüpfte Entität ({relation})"
    data[field] = entity
    colors[field] = "green"
    names[field] = []

    fields = [field] + fields

    return render_template("objekt_form_buttons.html",
                           fields=fields, data=data, colors=colors,
                           name_entities=names,
                           source_node_id=source, relation_name=relation)


# --------------------------------------------------------------
# OPTIONAL — keep "ensure_default_user" here or move to auth.py
# --------------------------------------------------------------
def ensure_default_user():
    with driver.session() as session:
        result = session.run("MATCH (u:User {username:$u}) RETURN u", {"u":"admin"})
        if not result.single():
            session.run(
                "CREATE (u:User {id:1, username:'admin', password_hash:$p})",
                {"p": generate_password_hash("secret")}
            )
