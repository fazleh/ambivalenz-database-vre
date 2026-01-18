from flask import Flask, jsonify, Markup, request, render_template, redirect, url_for, flash
from flask_login import LoginManager, login_user, login_required, logout_user, UserMixin
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from neo4j import GraphDatabase
import os, csv, re, random, subprocess
from openpyxl import Workbook, load_workbook
import spacy
from urllib.parse import unquote

# --- Configuration ---
number = random.randint(1, 100)
category = "article"
objekt_id = f"{category}_{number}"
DIR = "/home/melahi/code/A-mediawiki-project/neo4j-upload/dataset/german/input/"
CSV_FILE = os.path.join(DIR, f"entity_{objekt_id}.csv")
EXCEL_FILE = "objekt_data.xlsx"

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Load spaCy German model
nlp = spacy.load("de_core_news_sm")

# Ensure Excel file exists
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Objekt Informationen"
    ws.append(["Property", "Value", "Status", "Name Entity"])
    wb.save(EXCEL_FILE)


def get_field_list():
    """Read 'Property' column from Excel file, or use defaults."""
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        props = [str(row[0]).strip() for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]
        if props:
            return props
    return [
        "Titel", "Künstler*in/Autor*in", "Sichtbare oder genannte Personen",
        "Entstehungsjahr", "Ort der Entstehung / Nutzung", "Gattung / Genre",
        "Technik", "Dimensionen", "Kurzbeschreibung", "Motive / Topoi",
        "Narrative / Diskurse", "Historischer Kontext",
        "Antiziganistische / Stigmatisierende Elemente", "Agency", "Verknüpfung",
        "Narrativwandel bei Medienwechsel", "Rezeptionsweg", "Sammlung / Archiv",
        "Provenienz", "Literatur", "Ausstellungen / Aufführungen / Veröffentlichungen",
        "Objekt- oder Werkteil", "Rechte / Lizenzen", "Digitalisat-Link/Pfad",
        "Metadaten-Status", "Erfasst von", "Erfassungsdatum", "Kommentar / Anmerkung",
        "Versionsgeschichte"
    ]


def extract_name_entities(text):
    """Extract PER, ORG, LOC entities."""
    if not text:
        return []
    doc = nlp(text)
    seen, result = set(), []
    for ent in doc.ents:
        if ent.label_ in ("PER", "ORG", "LOC") and ent.text.strip() not in seen:
            seen.add(ent.text.strip())
            result.append(ent.text.strip())
    return result


@app.route('/')
def form():
    fields = get_field_list()
    data, colors, name_entities = {}, {}, {}
    return render_template("objekt_form_buttons.html", fields=fields, data=data, colors=colors, name_entities=name_entities)


@app.route('/upload', methods=['POST'])
def upload():
    file = request.files.get('file')
    if not file or file.filename == '':
        return "❌ Keine Datei ausgewählt", 400
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
    file.save(filepath)
    wb = load_workbook(filepath)
    ws = wb.active

    data, colors, name_entities = {}, {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        key, val = str(row[0]).strip(), str(row[1]).strip() if len(row) > 1 and row[1] else ""
        data[key] = val
        colors[key] = "yellow" if val else "red"
        name_entities[key] = extract_name_entities(val)

    return render_template("objekt_form_buttons.html", fields=list(data.keys()), data=data, colors=colors, name_entities=name_entities)


@app.route('/extract_entities', methods=['POST'])
def extract_entities():
    fields = get_field_list()
    data, colors, name_entities = {}, {}, {}
    for key in fields:
        val = request.form.get(key, "").strip()
        ne = extract_name_entities(val)
        data[key], name_entities[key] = val, ne
        colors[key] = "green" if ne else "yellow" if val else "red"
    return render_template("objekt_form_buttons.html", fields=fields, data=data, colors=colors, name_entities=name_entities)


@app.route('/submit', methods=['POST'])
def submit():
    posted = request.form
    rows = []

    # Iterate over all fields in form
    for key, value in posted.items():
        if key.startswith("custom_property_"):
            index = key.split("_")[-1]
            prop_name = value.strip()
            prop_val = posted.get(f"custom_value_{index}", "").strip() or "No data"
            prop_status = posted.get(f"traffic_custom_{index}", "red")
            if prop_name or prop_val:
                rows.append([prop_name or "No data", prop_val or "No data", prop_status, ""])
        elif key.startswith("traffic_") or key in ("source_node_id", "relation_name", "category", "nodeType"):
            continue  # skip control fields
        else:
            val = value.strip() or "No data"
            status = posted.get(f"traffic_{key}", "red")
            rows.append([key, val, status, ""])

    # Final row with nodeType / category
    node_type_value = posted.get("nodeType", posted.get("category", "article"))
    rows.append(["nodeType", node_type_value, "", ""])

    # Save CSV
    os.makedirs(os.path.dirname(CSV_FILE), exist_ok=True)
    with open(CSV_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Property", "text", "Status", "Name Entity"])
        writer.writerows(rows)

    # Run Java CREATE or RELATION
    source_node_id = posted.get("source_node_id")
    relation_name = posted.get("relation_name")
    try:
        if source_node_id and relation_name:
            cmd = ["java", "-jar", "/home/melahi/code/A-mediawiki-project/neo4j-upload/target/QuestionGrammarGenerator.jar", "RELATION"]
        else:
            cmd = ["java", "-jar", "/home/melahi/code/A-mediawiki-project/neo4j-upload/target/QuestionGrammarGenerator.jar", "CREATE"]
        result = subprocess.run(cmd, capture_output=True, text=True, check=True)
        print(result.stdout)
    except subprocess.CalledProcessError as e:
        print(e.stderr)
        return f"<h3>ERROR: {e.stderr}</h3>"

    return redirect(url_for('new_form'))


@app.route('/open_entity_form/<path:relation>/<path:source_node_id>/<path:entity_name>')
def open_entity_form(relation, source_node_id, entity_name):
    """Open new form showing which entity was clicked."""
    relation = unquote(relation)
    entity_name = unquote(entity_name)
    source_node_id = unquote(source_node_id)

    fields = get_field_list()
    data, colors, name_entities = {}, {}, {}

    special_field = f"Verknüpfte Entität ({relation})"
    data[special_field] = entity_name
    colors[special_field] = "green"
    name_entities[special_field] = []

    fields = [special_field] + fields
    return render_template("objekt_form_buttons.html", fields=fields, data=data, colors=colors, name_entities=name_entities,
                           source_node_id=source_node_id, relation_name=relation)


if __name__ == '__main__':
    app.run(debug=True, port=5010)
