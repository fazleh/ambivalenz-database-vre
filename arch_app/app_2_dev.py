# merged_app.py
from flask import Flask, jsonify, request, render_template, redirect, url_for, flash
from markupsafe import Markup

from markupsafe import Markup as MarkupType
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from neo4j import GraphDatabase
from openpyxl import Workbook, load_workbook
from urllib.parse import unquote
import os
import csv
import re
import random
import subprocess
import spacy

# -----------------------
# App configuration
# -----------------------
app = Flask(__name__, static_url_path="/static")
app.secret_key = os.getenv("FLASK_SECRET_KEY", "your-secret-key")
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# -----------------------
# Login manager
# -----------------------
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# -----------------------
# Neo4j connection
# -----------------------
driver = GraphDatabase.driver(
    os.getenv("NEO4J_URI", "bolt://neo4j:7687"),
    auth=(os.getenv("NEO4J_USER", "neo4j"), os.getenv("NEO4J_PASSWORD", "password"))
)

# -----------------------
# spaCy (German) for NER
# -----------------------
# Make sure you have installed the model: python -m spacy download de_core_news_sm
nlp = spacy.load("de_core_news_sm")

# -----------------------
# Excel/CSV config for /add_data
# -----------------------
EXCEL_FILE = "objekt_data.xlsx"
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Objekt Informationen"
    ws.append(["Property", "Value", "Status", "Name Entity"])
    wb.save(EXCEL_FILE)

# -----------------------
# User model (single unified)
# -----------------------
class User(UserMixin):
    def __init__(self, id_, username, password_hash, name=None, institution=None, user_group=None):
        self.id = id_
        self.username = username
        self.password_hash = password_hash
        self.name = name
        self.institution = institution
        self.user_group = user_group

@login_manager.user_loader
def load_user(user_id):
    with driver.session() as session:
        result = session.run(
            "MATCH (u:User) WHERE u.id = $id RETURN u", {"id": int(user_id)}
        )
        record = result.single()
        if record:
            u = record["u"]
            # some records may not have all fields
            return User(
                u.get("id"), u.get("username"), u.get("password_hash"),
                u.get("name"), u.get("institution"), u.get("user_group")
            )
    return None

# -----------------------
# Utility functions
# -----------------------
def get_nodes():
    with driver.session() as session:
        result = session.run("MATCH (n:Person) RETURN n.name AS name LIMIT 10")
        return [record["name"] for record in result]

def run_cypher_query(cypher_query):
    with driver.session() as session:
        try:
            result = session.run(cypher_query)
            keys = result.keys()
            records = [record.data() for record in result]
            return keys, records, None
        except Exception as e:
            return [], [], str(e)

def is_valid_password(password):
    return (
        re.search(r'[A-Z]', password) and
        re.search(r'[a-z]', password) and
        re.search(r'[!@#$%^&*()_\-.,?":{}|<>]', password)
    )

def is_valid_email(email):
    email_regex = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return re.match(email_regex, email)

def get_universities_from_csv():
    filepath='open-data/universities.csv'
    try:
        with open(filepath, newline='', encoding='utf-8') as csvfile:
            return [row[0] for row in csv.reader(csvfile)]
    except FileNotFoundError:
        return []

def extract_name_entities(text):
    """Return list of unique name entities (PER, ORG, LOC) from text using spaCy."""
    if not text:
        return []
    doc = nlp(text)
    seen = set()
    ents = []
    for ent in doc.ents:
        if ent.label_ in ("PER", "ORG", "LOC"):
            txt = ent.text.strip()
            if txt not in seen:
                seen.add(txt)
                ents.append(txt)
    return ents

def get_field_list():
    """Read 'Property' column from Excel file, or use defaults."""
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        props = [str(row[0]).strip() for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]
        if props:
            return props
    return [
        "Titel", "Künstler*in/Autor*in", "Sichtbare oder genannte Personen",
        "Entstehungsjahr", "Ort der Entstehung / Nutzung", "Gattung / Genre",
        "Technik", "Dimensionen", "Kurzbeschreibung", "Motive / Topoi",
        "Narrative / Diskurse", "Historischer Kontext",
        "Antiziganistische / Stigmatisierende Elemente", "Agency", "Verknüpfung",
        "Narrativwandel bei Medienwechsel", "Rezeptionsweg", "Sammlung / Archiv",
        "Provenienz", "Literatur", "Ausstellungen / Aufführungen / Veröffentlichungen",
        "Objekt- oder Werkteil", "Rechte / Lizenzen", "Digitalisat-Link/Pfad",
        "Metadaten-Status", "Erfasst von", "Erfassungsdatum", "Kommentar / Anmerkung",
        "Versionsgeschichte"
    ]

# -----------------------
# Routes from the FIRST script (kept intact)
# -----------------------

@app.route("/api/nodes")
def nodes():
    data = get_nodes()
    return jsonify(data)

@app.route("/sections/")
def sections():
    return render_template("sections.html")

@app.route("/visual_art", methods=["GET", "POST"])
def visual_art():
    keys, records, error, submitted_query = [], [], None, ""
    if request.method == "POST":
        submitted_query = request.form.get("cypher_query")
        keys, records, error = run_cypher_query(submitted_query)
    return render_template("visual_art.html", keys=keys, records=records, error=error, query=submitted_query)

@app.route("/audio")
def audio():
    return render_template("audio.html")

@app.route("/video")
def video():
    return render_template("video.html")

@app.route("/book", methods=["GET", "POST"])
def book():
    keys, records, error, submitted_query = [], [], None, ""
    if request.method == "POST":
        submitted_query = request.form.get("cypher_query")
        keys, records, error = run_cypher_query(submitted_query)
    return render_template("book.html", keys=keys, records=records, error=error, query=submitted_query)

@app.route("/article", methods=["GET", "POST"])
def article():
    keys, records, error, submitted_query = [], [], None, ""
    if request.method == "POST":
        submitted_query = request.form.get("cypher_query")
        keys, records, error = run_cypher_query(submitted_query)
    return render_template("article.html", keys=keys, records=records, error=error, query=submitted_query)

@app.route("/poem", methods=["GET", "POST"])
def poem():
    keys, records, error, submitted_query = [], [], None, ""
    if request.method == "POST":
        submitted_query = request.form.get("cypher_query")
        keys, records, error = run_cypher_query(submitted_query)
    return render_template("poem.html", keys=keys, records=records, error=error, query=submitted_query)

@app.route("/song", methods=["GET", "POST"])
def song():
    keys, records, error, submitted_query = [], [], None, ""
    if request.method == "POST":
        submitted_query = request.form.get("cypher_query")
        keys, records, error = run_cypher_query(submitted_query)
    return render_template("song.html", keys=keys, records=records, error=error, query=submitted_query)

@app.route("/legal_text", methods=["GET", "POST"])
def legal_text():
    keys, records, error, submitted_query = [], [], None, ""
    if request.method == "POST":
        submitted_query = request.form.get("cypher_query")
        keys, records, error = run_cypher_query(submitted_query)
    return render_template("legal_text.html", keys=keys, records=records, error=error, query=submitted_query)

@app.route("/Person", methods=["GET", "POST"])
def Person():
    keys, records, error, submitted_query = [], [], None, ""
    if request.method == "POST":
        submitted_query = request.form.get("cypher_query")
        keys, records, error = run_cypher_query(submitted_query)
    return render_template("Person.html", keys=keys, records=records, error=error, query=submitted_query)

@app.route("/poster", methods=["GET", "POST"])
def poster():
    keys, records, error, submitted_query = [], [], None, ""
    if request.method == "POST":
        submitted_query = request.form.get("cypher_query")
        keys, records, error = run_cypher_query(submitted_query)
    return render_template("poster.html", keys=keys, records=records, error=error, query=submitted_query)

@app.route("/portrait", methods=["GET", "POST"])
def portrait():
    keys, records, error, submitted_query = [], [], None, ""
    if request.method == "POST":
        submitted_query = request.form.get("cypher_query")
        keys, records, error = run_cypher_query(submitted_query)
    return render_template("portrait.html", keys=keys, records=records, error=error, query=submitted_query)

@app.route("/about/")
def about():
    return render_template("about.html")

@app.route("/context-project/")
def context_project():
    return render_template("context_project.html")

@app.route("/history-of-romarchive/")
def history_of_romarchive():
    return render_template("history_of_romarchive.html")

@app.route("/curators/")
def curators():
    return render_template("curators.html")

@app.route("/ethical-guidelines/")
def ethical_guidelines():
    return render_template("ethical_guidelines.html")

@app.route("/collection-policy/")
def collection_policy():
    return render_template("collection_policy.html")

@app.route("/faq/")
def faq():
    return render_template("faq.html")

@app.route("/search/")
def search():
    return render_template("search.html")

@app.route("/terms/")
def terms():
    return render_template("terms.html")

@app.route("/contact/")
def contact():
    return render_template("contact.html")

@app.route("/collection/politics-of-photography/")
def politics_of_photography():
    return render_template("politics_of_photography.html")

@app.route("/visual_art/Individual_page")
def Individual_page():
    return render_template("Individual_page.html")

@app.route("/visual_art/Individual_page_var")
def Individual_page_var():
    section = request.args.get("section", default=None)

    title = ""
    artist_info = """
                   Das Bild ist ambivalent: Es zeigt einerseits Respekt für die Ästhetik und „Malerhaftigkeit“ 
                   der dargestellten Menschen, andererseits reproduziert es stereotype und exotisierende Merkmale. 
                   Es kann sowohl als bewundernde Darstellung als auch als visuelle Festschreibung von „Andersartigkeit“ gelesen werden.
                """
    image_path = "private/Zwei_Zigeuner.png"

    content = ""
    if section == "Objekt_Informationen":
        return render_template("Individual_page_var.html")
    elif section == "Inhaltliche_Beschreibung":
        try:
            with open("s1-templates/zwei_zigeuner_inh_besc.html", "r", encoding="utf-8") as file:
                table_html = file.read()
            content = Markup(table_html)
        except FileNotFoundError:
            content = "Table file not found."
    elif section == "Semantische_Annotation":
        try:
            with open("s1-templates/zwei_zigeuner_sem_ann.html", "r", encoding="utf-8") as file:
                table_html = file.read()
            content = Markup(table_html)
        except FileNotFoundError:
            content = "Table file not found."
    elif section == "Semantische_Relationen":
        try:
            with open("s1-templates/zwei_zigeuner_sem_rel.html", "r", encoding="utf-8") as file:
                table_html = file.read()
            content = Markup(table_html)
        except FileNotFoundError:
            content = "Table file not found."
    elif section == "Technische_rechtliche":
        try:
            with open("s1-templates/zwei_zigeuner_tech_rech.html", "r", encoding="utf-8") as file:
                table_html = file.read()
            content = Markup(table_html)
        except FileNotFoundError:
            content = "Table file not found."
    else:
        return render_template("Individual_page_var.html")

    return render_template(
        "Individual_page_var.html",
        title=title,
        artist_info=artist_info,
        image_path=image_path,
        section=section,
        content=content,
    )

def fetch_object_by_name(name):
    with driver.session() as session:
        result = session.run("MATCH (o:Objekt {name: $name}) RETURN o", name=name)
        record = result.single()
        if record:
            return dict(record["o"])
        return None

@app.route("/book/Roma_Sinti")
def Roma_Sinti():
    title = "Roma & Sinti : Zigeuner-Darstellungen der Moderne"
    artist_info = """
    Entstehung: frühes 20. Jahrhundert, kurz vor dem Ersten Weltkrieg Ungarn als Teil der Donaumonarchie 
    mit starken ethnischen Spannungen Zunehmende politische Kontrolle und Militarisierung (z. B. Haarschnitt 
    der Rekruten ab 1914 als Disziplinierungsmaßnahme) Stigmatisierung und Marginalisierung von Roma-Gruppen 
    im gesamten europäischen Raum Beginn einer staatlich regulierten Rassifizierung  Ermittlung der Daten…
    """
    image_path = "private/Roma_Sinti.png"
    try:
        with open("s1-templates/book_roma_all.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/book/Zigeuner_Bukarest")
def Zigeuner_Bukarest():
    title = "Zigeuner auf Baustellen in Bukarest"
    artist_info = """
    Der Hunsrücker Maler verdient sein Geld mit Malerarbeiten in Neubauten und beschreibt die Zustände und Rangordnung +auf Baustellen in Bukarest. Auffällig lang ist seine Charakterisierung der Roma
    """
    image_path = "private/Zigeuner_Bukarest.png"
    try:
        with open("s1-templates/book_Zigeuner_Bukarest.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/legal_text/Gewerbeordnung_1883")
def Gewerbeordnung_1883():
    title = "Gewerbeordnung 1883 – Ausführungsbestimmungen"
    artist_info = """
    Bekanntmachung des  Handels- und Gewerbewesens vom 31. Oktober 1883 betreffend Ausführungsbestimmungen zur Gewerbeordnung für das Deutsche Reich (Reichs-Gesetzblatt 1883 Seite 177)
    """
    image_path = "private/Gewerbeordnung_1883.png"
    try:
        with open("s1-templates/book_gewerbeordnung_1883_all.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/legal_text/Gewerbeordnung_1904")
def Gewerbeordnung_1904():
    title = "Preußische Ausführungsanordnung zur Gewerbeordnung des Deutschen Reichs vom 1. Mai 1904"
    artist_info = """
    Versagen von Wandergewerbescheinen bei inländischen Personen, die als "Zigeuner" gelabelt werden; Markieren von Wandergewerbescheinen mit dem Zustaz "Zigeuner" oder wenn Eigenschaft nicht fesstellbar mit dem Zusatz "Zieht nach Zigeunerart im Land umher"
    """
    image_path = "private/Gewerbeordnung_1904.png"
    try:
        with open("templates/legal_text_gewerbeordnung_1904.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/portrait/portrait_flower_1")
def portrait_flower_1():
    title = "Zigeunerinnen mit Sonnenblumen"
    artist_info = """
    Expressionistische Darstellung dreier Figuren in einer ländlichen Szene. Es handelt sich um zwei Frauen und ein Baby. Figuren und Objekte sind teils stark vereinfacht dargestellt. 
    Im Zentrum steht eine aufrecht stehende Frau mit dunkler Haut, kurzem schwarzen Haar und einem langen, hellen, mit Punkten besetzten Gewand. Die Hände hält sie vor der Brust verschränkt. Rechts unten sitzt eine zweite, ebenfalls dunkelhäutige Frau mit leuchtend gelbem Kopftuch und Oberteil und einem braun-blau gestreiftem Rock. Sie stillt ein Baby.
    Links neben der stehenden Frau ragt eine große Sonnenblume ins Bild. Im Hintergrund sieht man weiß verputzte Häuser mit orangefarbenen Dächern, grüne Bäume und ein Pferd.
    """
    image_path = "private/portrait_flower_1.jpeg"
    try:
        with open("s1-templates/portrait_flower_1.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/portrait/portrait_person_2")
def portrait_person_2():
    title = "Zigeuner"
    artist_info = """
    Porträt eines jungen Mannes. Er trägt bürgerliche Kleidung: kariertes Jackett mit einem gelb-bräunlichen Karomuster auf einem dunkelbraunen Grundton, weißes Hemd und blaue Krawatte. Auf der linken Seite des Jacketts befindet sich ein Flicken. Der Mann hat dunkles, glänzendes, lockiges Haar, das rechts gescheitelt ist. Sein Blick ist noch vorne gerichtet. Er hat buschige Augenbraun und einen angedeuteten Schnurrbart. Der Hintergrund des Bilds ist grau. 
    """
    image_path = "private/portrait_person_2.jpeg"
    try:
        with open("s1-templates/portrait_person_2.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/poster/Weltausstellung_Paris_1900_1")
def Weltausstellung_Paris_1900_1():
    title = "Exposition 1900/ 1900/ L'andalousie/ au/ temps des Maures/ ADMINISTRATION 74 Bd. HAUSSMANN, PARIS 8 "
    artist_info = """
    Überlebensgroßes Werbeplakat für die Ausstellung L'Andalousie au temps des Maures (Andalusien zur Zeit der Mauren) auf der Weltausstellung 1900 in Paris.
    Inhaltliche Beschreibung laut Archiv: "Gitane lisant dans la main d'une Espagnole, groupe d'Espagnoles, décor d'inspiration mauresque" (Gitane, die einer Spanierin aus der Hand liest; Gruppe spanischer Frauen; maurisch inspiriertes Dekor).
    Das Plakat kündigt in Figur der alten Handleserin das Unterhaltungsangebot der Wahrsagerei im Programm der Ausstellung an. Zugleich wirbt das Plakat für spanische Tanz-Auftritte in Figur der jungen Spanierin mit den Attributen Manila-Tuch, Volantrock, Fächer und Blume im Haar. Der Gegensatz zwischen der jungen und der alten Frau im Kontext der Wahrsagerei ruft die Figur der La Celestina, der legendären spanischen Kupplerin aus Fernando de Rojas gleichnamiger Tragikomödie (1499), auf. Am unteren Plakatrand ist die Adresse der französischen Betreibergesellschaft der Ausstellung angegeben. Das Plakat ist eines von mehreren Werbeplakaten für die Ausstellung, die von Mitgliedern der Société des Peintres Orientalistes Français gestaltet wurden.
    """
    image_path = "private/TP2_Weltausstellung_Paris_1900_1.jpg"
    try:
        with open("s1-templates/poster_Weltausstellung_Paris_1900_1.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/book/Weltausstellung_Paris_1900_2")
def Weltausstellung_Paris_1900_2():
    title = "L’ANDALOUSIE AU TEMPS DES MAURES – LES GITANES  "
    artist_info = """
    L'Andalousie» , qui occupe au Trocadéro plus de cinq mille mètres, est une des attractions les plus importantes de l'Exposition. Nous voici, après quelques pas, transportés en pleine Espagne; à droite, de vieilles maisons romanes de la province de Tolède; à gauche, peinte en trompe-l'oeil, une vue panoramique de l'Alhambra de Grenade avec le Monte Sacro, refuge des gitanes, aux danses endiablées et aux déhanchements si provocants.  
    """
    image_path = "private/TP2_Weltausstellung_Paris_1900_2.jpg"
    try:
        with open("s1-templates/poster_Weltausstellung_Paris_1900_2.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/song/Nordungarn_Q1_TH")
def Nordungarn_Q1_TH():
    title = "Zigeunertaufe in Nordungarn"
    artist_info = """
    Beschreibung von Ritualen der Roma die vor der klassischen Kirchlichen Taufe, von Familie, Verwandten, Paten  und in der ganzen Gemeinschaft abgehlten werden. 
    """
    image_path = "private/Nordungarn_Q1_TH.png"
    try:
        with open("s1-templates/song_Nordungarn_Q1_TH.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/article/article_Zigeunern_Q2_TH")
def article_Zigeunern_Q2_TH():
    title = "Der Seelenloskauf bei den mohammedanischen Zigeunern der Balkanländer "
    artist_info = """
         Beschreibung von Totenritualen und Eheritualen von moslimischen Seeshaften und "Wanderzigeunern    
         """
    image_path = "private/Zigeunern_Q2_TH.png"
    try:
        with open("s1-templates/article_Zigeunern_Q2_TH.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/article/Zigeunerisch_Q3_TH")
def Zigeunerisch_Q3_TH():
    title = "Zigeunerisch "
    artist_info = """
         Das Gedicht behandelt den Jahreszeitenwechsel und den Alltags, Liebes-, Arbeitsleben der Roma. Es will die Lebensweise der "Zeltzigeuner" romantisch und tragisch wieder geben. In ihm  tauchen die Beziehungen zu Leben und Tod und die Lebensbedingungen der Roma auf und was mit dem Gedicht als typisch "Zigeunerisch" hervorgehoben wird, wie der Titel es schon vermeldet.     
         """
    image_path = "private/Zigeunerisch_Q3_TH.png"
    try:
        with open("s1-templates/poem-Zigeunerisch_Q3_TH.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/visual_art/Zwei_Zigeuner")
def Zwei_Zigeuner():
    title = "Zwei Zigeuner"
    artist_info = """
    Das Bild Bild ist ambivalent: Es zeigt einerseits Respekt für die Ästhetik und „Malerhaftigkeit“ der dargestellten Menschen, andererseits reproduziert es 
    stereotype und exotisierende Merkmale.Es kann sowohl als bewundernde Darstellung als auch als visuelle Festschreibung von „Andersartigkeit“ gelesen werden.
    """
    image_path = "private/Zwei_Zigeuner.png"
    try:
        with open("s1-templates/zwei_zigeuner_obj_info_all.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/visual_art/Ilonka")
def Ilonka():
    title = "Ilonka"
    artist_info = """
    Fremdbild vs. Selbstpräsenz: Ilonka wird mit Attributen der „Zigeunerin“ ausgestattet: dunkle Kleidung, Goldschmuck, 
    sinnlicher Blick Deutungsrahmen: Exotisierung – aber: Sie schaut selbstbewusst, konfrontativ zurück
    """
    image_path = "private/Ilonka.png"
    try:
        with open("s1-templates/painting_ilonka_all.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/visual_art/Zigeuner")
def Zigeuner():
    title = "Pipázó cigány (Pfeife rauchender „Zigeuner“)"
    artist_info = (
        "Im Gegensatz zu vielen Darstellungen dieser Zeit: keine offensichtlichen antiziganistischen Stereotype. "
        "Der dargestellte Mann wird nicht exotisiert oder kriminalisiert, sondern erscheint würdevoll, selbstbewusst und bürgerlich. "
        "Titel „Pipázó cigány“ (übersetzt: „Pfeife rauchender Zigeuner“) verwendet die historisch belastete Bezeichnung „Zigeuner“. "
        "Schon durch die Benennung wird eine Fremdzuschreibung vorgenommen: Statt den individuellen Namen des Modells zu nennen, wird seine ethnische Zugehörigkeit betont und stereotyp markiert. "
        "Die Wortwahl verstärkt eine folkloristische Rahmung („Zigeuner“ + Pfeife als romantisierende, exotisierende Attribute). Auch wenn das Bild selbst individuelle Würde zeigt, reproduziert der Titel eine kulturelle Distanz und eine Fremddefinition."
    )
    image_path = "private/Zigeuner.png"
    try:
        with open("s1-templates/painting_zigeuner_all.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/visual_art/Katze")
def Katze():
    title = "Zwei Zigeunerin** mit Katze*"
    artist_info = (
        "•Rassistischer Werktitel (Begriff „Zigeunerinnen“). "
        "Darstellung als exotisch, erotisch, unzivilisiert. "
        "Fetischisierung weiblicher Romnja-Körper. "
        "Reproduktion kolonialer Zuschreibungen („das Andere“)"
    )
    image_path = "private/Katze.png"
    try:
        with open("s1-templates/painting_katze_all.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/visual_art/Zigeunerin")
def Zigeunerin():
    title = "Zigeunerin"
    artist_info = """
        Eine Frau posiert Frau braunem Hintergrund und Vorhang links vor dem ein Schemel über den ein buntes Gewand geworfen ist. Sie dreht den Rücken zum Betrachter, während ihr Gesicht  im Profil zu sehen ist und ihr Arm unter einem grünen übergeschlagenen Tuch in die Seite gestützt ist. Darunter trägt sie ein rosa Kleid mit weißen Partien.   
        """
    image_path = "private/Zigeunerin.png"
    try:
        with open("s1-templates/painting_Zigeunerin_all.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)


app = Flask(__name__)

# ------------------------------------------------------
# Neo4j connection
# ------------------------------------------------------
driver = GraphDatabase.driver(
    "bolt://neo4j:7687",
    auth=("neo4j", "password")
)

# ------------------------------------------------------
# Fetch all paintings (nodeType = "painting")
# ------------------------------------------------------
def fetch_all_paintings():
    """
    Returns a list of nodes where nodeType = 'painting'
    """
    with driver.session() as session:
        query = """
        MATCH (n)
        WHERE n.nodeType = "painting"
        RETURN n
        """
        results = session.run(query)
        return [dict(record["n"]) for record in results]


# ------------------------------------------------------
# Fetch a single painting by name
# ------------------------------------------------------
def fetch_node_data(node_name):
    with driver.session() as session:
        query = """
        MATCH (n {name: $name})
        RETURN n
        """
        result = session.run(query, name=node_name)
        record = result.single()

    if record:
        return dict(record["n"])
    return None


# ------------------------------------------------------
# Convert node data → HTML table
# ------------------------------------------------------
def generate_html_table(node_data):
    html = "<table border='1'><tr><th>Property</th><th>Value</th></tr>"
    for key, value in node_data.items():
        html += f"<tr><td>{key}</td><td>{value}</td></tr>"
    html += "</table>"
    return Markup(html)


# ------------------------------------------------------
# Automatically create a Flask route for one painting
# ------------------------------------------------------
def create_painting_route(painting_name):

    route = f"/visual_art/{painting_name}"

    @app.route(route)
    def painting_page(painting_name=painting_name):

        node_data = fetch_node_data(painting_name)

        if not node_data:
            return f"<h2>No Neo4j data found for: {painting_name}</h2>"

        title = node_data.get("Titel", painting_name)
        artist_info = node_data.get("Kurzbeschreibung", "")
        image_path = node_data.get("image_path", "")

        content = generate_html_table(node_data)

        return render_template(
            "Individual_page_var.html",
            title=title,
            artist_info=artist_info,
            image_path=image_path,
            content=content
        )


# ------------------------------------------------------
# ON APPLICATION START:
# Load all paintings and generate routes
# ------------------------------------------------------
all_paintings = fetch_all_paintings()

for painting in all_paintings:
    if "name" in painting:
        create_painting_route(painting["name"])

print(f"✔ Loaded {len(all_paintings)} painting routes from Neo4j.")


@app.route("/Person/Person_Friedrich")
def Person_Friedrich():
    title = "Ströher, Friedrich Karl"
    artist_info = """
    Olga auf grüner Wiese, 1910 | Brücke in Paris mit Omnibus, 1904 | Spanierin beim Nähen, 1912, Kornerde vor Irmenach, 1924 | Gefallenen-Ehrenmale in Irmenach und Hirschfeld (letzteres 1933 als unheroisch abgebaut) | Veröffentlichungen: Erinnerungen 1876-1911, MS | Reise nach Südfrankreich 1910, Manuskript | Irmenach 1923/25 Irmenach - in: Archiv für bildende Kunst, Germanisches Nationalmuseum Nürnberg | Privater Nachlaß: Peter Ströher (Sohn) | Sammlung Ströher (Dauerleihgabe) im Hunsrück-Museum, Schloß Simmern
        """
    image_path = "private/Friedrich.png"
    try:
        with open("s1-templates/Person_Friedrich.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/Person/Person_Kampf")
def Person_Kampf():
    title = "Prof. Dr. Arthur Kampf"
    artist_info = ""
    image_path = "private/Kampf.png"
    try:
        with open("s1-templates/Person_Kampf.html", "r", encoding="utf-8") as file:
            table_html = file.read()
        content = Markup(table_html)
    except FileNotFoundError:
        content = "Table file not found."
    return render_template("Individual_page_var.html", title=title, artist_info=artist_info, image_path=image_path, content=content)

@app.route("/index/Metadata/")
def Metadata():
    title = "Metadata Description"
    properties_1 = []
    try:
        with open('data/newdata/description_1.csv', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                properties_1.append((row.get("Property"), row.get("text")))
    except FileNotFoundError:
        properties_1 = []

    properties_2 = []
    try:
        with open('data/newdata/description_2.csv', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                properties_2.append((row.get("Property"), row.get("text")))
    except FileNotFoundError:
        properties_2 = []

    properties_3 = []
    try:
        with open('data/newdata/description_3.csv', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                properties_3.append((row.get("Property"), row.get("text")))
    except FileNotFoundError:
        properties_3 = []

    properties_4 = []
    try:
        with open('data/newdata/description_4.csv', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                properties_4.append((row.get("Property"), row.get("text")))
    except FileNotFoundError:
        properties_4 = []

    properties_5 = []
    try:
        with open('data/newdata/description_5.csv', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                properties_5.append((row.get("Property"), row.get("text")))
    except FileNotFoundError:
        properties_5 = []

    return render_template(
        "Metadata_page_var.html",
        title=title,
        properties_1=properties_1, properties_2=properties_2,
        properties_3=properties_3, properties_4=properties_4,
        properties_5=properties_5
    )

@app.route("/supporters/")
def supporters():
    return "<h1>Supporters Page</h1>"

@app.route("/imprint/")
def imprint():
    return "<h1>Imprint Page</h1>"

@app.route("/privacy/")
def privacy():
    return "<h1>Privacy Statement</h1>"

@app.route("/project-detail")
def project_detail():
    return "Welcome to the Project Detail page!"

@app.route('/delete_all_data')
def delete_all_data():
    with driver.session() as session:
        session.run("MATCH (n) DETACH DELETE n")
    return "All data deleted from Neo4j."

# -----------------------
# Auth / register / login / logout
# -----------------------
@app.route('/register', methods=['GET', 'POST'])
def register():
    universities = get_universities_from_csv()
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        email = request.form.get('email', '')

        if not is_valid_email(email):
            flash('Please enter a valid email address.')
            return redirect(url_for('register'))

        if not is_valid_password(password):
            flash('Password must include at least one uppercase letter, one lowercase letter, and one symbol !@#$%^&*()_-.,?":{}|<>')
            return redirect(url_for('register'))

        name = request.form.get('name')
        institution = request.form.get('institution')
        user_group = request.form.get('user_group')
        password_hash = generate_password_hash(password)

        with driver.session() as session:
            existing_user = session.run("MATCH (u:User {username: $username}) RETURN u", {"username": username}).single()
            existing_email = session.run("MATCH (u:User {email: $email}) RETURN u", {"email": email}).single()

            if existing_user:
                flash("Username already taken.")
                return redirect(url_for('register'))
            if existing_email:
                flash("Email already registered.")
                return redirect(url_for('register'))

            session.run(
                """
                CREATE (u:User {
                    id: toInteger(timestamp()),
                    username: $username,
                    email: $email,
                    password_hash: $password_hash,
                    name: $name,
                    institution: $institution,
                    user_group: $user_group
                })
                """,
                {
                    "username": username,
                    "email": email,
                    "password_hash": password_hash,
                    "name": name,
                    "institution": institution,
                    "user_group": user_group
                }
            )

            flash("Registration successful! You can now log in.")
            return redirect(url_for('login'))

    return render_template('register.html', universities=universities)

# Root/login route(s)
@app.route('/', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')

        with driver.session() as session:
            result = session.run(
                "MATCH (u:User {username: $username}) RETURN u",
                {"username": username}
            )
            record = result.single()
            if record:
                u = record["u"]
                if check_password_hash(u["password_hash"], password):
                    user = User(u.get("id"), u.get("username"), u.get("password_hash"), u.get("name"), u.get("institution"), u.get("user_group"))
                    login_user(user)
                    return redirect(url_for('index'))

        flash("Invalid username or password")
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route("/index")
def index():
    return render_template("index.html")

# -----------------------
# add_data routes (SECOND script integrated under /add_data)
# -----------------------

@app.route('/add_data', methods=['GET'])
@login_required
def add_data():
    fields = get_field_list()
    data, colors, name_entities = {}, {}, {}
    return render_template("objekt_form_buttons.html", fields=fields, data=data, colors=colors, name_entities=name_entities)

@app.route('/add_data/upload', methods=['POST'])
def add_data_upload():
    file = request.files.get('file')
    if not file or file.filename == '':
        return "❌ No file selected", 400
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
    file.save(filepath)
    wb = load_workbook(filepath)
    ws = wb.active

    data, colors, name_entities = {}, {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        key = str(row[0]).strip()
        val = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        data[key] = val
        colors[key] = "yellow" if val else "red"
        name_entities[key] = extract_name_entities(val)

    return render_template("objekt_form_buttons.html", fields=list(data.keys()), data=data, colors=colors, name_entities=name_entities)

@app.route('/add_data/extract_entities', methods=['POST'])
def add_data_extract_entities():
    fields = get_field_list()
    data, colors, name_entities = {}, {}, {}
    for key in fields:
        val = request.form.get(key, "").strip()
        ne = extract_name_entities(val)
        data[key], name_entities[key] = val, ne
        colors[key] = "green" if ne else "yellow" if val else "red"
    return render_template("objekt_form_buttons.html", fields=fields, data=data, colors=colors, name_entities=name_entities)

@app.route('/add_data/submit', methods=['POST'])
def add_data_submit():
    posted = request.form
    rows = []

    # build rows from the posted form
    for key, value in posted.items():
        if key.startswith("custom_property_"):
            index = key.split("_")[-1]
            prop_name = value.strip()
            prop_val = posted.get(f"custom_value_{index}", "").strip() or "No data"
            prop_status = posted.get(f"traffic_custom_{index}", "red")
            if prop_name or prop_val:
                rows.append([prop_name or "No data", prop_val or "No data", prop_status, ""])
        elif key.startswith("traffic_") or key in ("source_node_id", "relation_name", "category", "nodeType"):
            continue  # skip control fields
        else:
            val = value.strip() or "No data"
            status = posted.get(f"traffic_{key}", "red")
            rows.append([key, val, status, ""])

    # final nodeType row
    node_type_value = posted.get("nodeType", posted.get("category", "article"))
    rows.append(["nodeType", node_type_value, "", ""])

    # Build CSV-like string (instead of writing real CSV)
    header = ["Property", "text", "Status", "Name Entity"]
    result_string = "=".join(header) + "\n"

    for row in rows:
        result_string += "=".join(str(col) for col in row) + "\n"
    # Generate unique object ID
    number = random.randint(1, 100)
    category = posted.get("category", "article")
    objekt_id = f"{category}_{number}"

    CSV_DIR = "/app/marburg-project/dataset/german/input/"
    JAR_PATH = "/app/marburg-project/target/QuestionGrammarGenerator.jar"

    source_node_id = posted.get("source_node_id")
    relation_name = posted.get("relation_name")
    # menu = "RELATION" if (source_node_id and relation_name) else "CREATE"

    menu ="CREATE_FROM_STRING"
    neo4j_uri = os.environ.get("NEO4J_URI", "bolt://neo4j:7687")
    neo4j_user = os.environ.get("NEO4J_USER", "neo4j")
    neo4j_pass = os.environ.get("NEO4J_PASSWORD", "password")
    javaflag="FALSE"

    cmd = [
        "java", "-jar", JAR_PATH,
        menu,
        CSV_DIR,  # directory containing generated text file
        neo4j_uri,
        neo4j_user,
        neo4j_pass,
        result_string,
        javaflag
    ]

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, check=True)
        print("stdout:", result.stdout)
        print("stderr:", result.stderr)
    except subprocess.CalledProcessError as e:
        print(e.stderr)
        return f"<h3>ERROR: {e.stderr}</h3>"

    return redirect(url_for('add_data'))

@app.route('/add_data/open_entity_form')
@login_required
def add_data_open_entity_form():
    relation = request.args.get("relation")
    source_node_id = request.args.get("source_node_id")
    entity_name = request.args.get("entity_name")

    fields = get_field_list()
    data, colors, name_entities = {}, {}, {}

    special_field = f"Verknüpfte Entität ({relation})"
    data[special_field] = entity_name
    colors[special_field] = "green"
    name_entities[special_field] = []

    fields = [special_field] + fields
    return render_template("objekt_form_buttons.html", fields=fields, data=data, colors=colors, name_entities=name_entities, source_node_id=source_node_id, relation_name=relation)

# -----------------------
# Optional: ensure default admin user
# -----------------------
def ensure_default_user():
    with driver.session() as session:
        result = session.run("MATCH (u:User {username: $username}) RETURN u", {"username": "admin"})
        if not result.single():
            hashed = generate_password_hash("secret")
            session.run(
                "CREATE (u:User {id: $id, username: $username, password_hash: $password_hash})",
                {"id": 1, "username": "admin", "password_hash": hashed}
            )

# -----------------------
# Run app
# -----------------------
if __name__ == "__main__":
    ensure_default_user()
    app.run(host='0.0.0.0', port=5000, debug=True)
