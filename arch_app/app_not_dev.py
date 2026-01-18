# merged_app.py
from flask import Flask, jsonify, request, render_template, redirect, url_for, flash
from markupsafe import Markup

from markupsafe import Markup as MarkupType
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from neo4j import GraphDatabase
from openpyxl import Workbook, load_workbook
from urllib.parse import unquote
import os
import csv
import re
import random
import subprocess
import spacy
from flask import Flask, render_template, request
from markupsafe import Markup
from neo4j import GraphDatabase

# CALL db.labels();
#
# -----------------------
# App configuration
# -----------------------
app = Flask(__name__, static_url_path="/static")
app.secret_key = os.getenv("FLASK_SECRET_KEY", "your-secret-key")
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# -----------------------
# Login manager
# -----------------------
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# -----------------------
# Neo4j connection
# -----------------------
driver = GraphDatabase.driver(
    os.getenv("NEO4J_URI", "bolt://neo4j:7687"),
    auth=(os.getenv("NEO4J_USER", "neo4j"), os.getenv("NEO4J_PASSWORD", "password"))
)

# -----------------------
# spaCy (German) for NER
# -----------------------
# Make sure you have installed the model: python -m spacy download de_core_news_sm
nlp = spacy.load("de_core_news_sm")

# -----------------------
# Excel/CSV config for /add_data
# -----------------------
EXCEL_FILE = "objekt_data.xlsx"
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Objekt Informationen"
    ws.append(["Property", "Value", "Status", "Name Entity"])
    wb.save(EXCEL_FILE)

# -----------------------
# User model (single unified)
# -----------------------
class User(UserMixin):
    def __init__(self, id_, username, password_hash, name=None, institution=None, user_group=None):
        self.id = id_
        self.username = username
        self.password_hash = password_hash
        self.name = name
        self.institution = institution
        self.user_group = user_group

@login_manager.user_loader
def load_user(user_id):
    with driver.session() as session:
        result = session.run(
            "MATCH (u:User) WHERE u.id = $id RETURN u", {"id": int(user_id)}
        )
        record = result.single()
        if record:
            u = record["u"]
            # some records may not have all fields
            return User(
                u.get("id"), u.get("username"), u.get("password_hash"),
                u.get("name"), u.get("institution"), u.get("user_group")
            )
    return None

# -----------------------
# Utility functions
# -----------------------
def get_nodes():
    with driver.session() as session:
        result = session.run("MATCH (n:Person) RETURN n.name AS name LIMIT 10")
        return [record["name"] for record in result]

def run_cypher_query(cypher_query):
    with driver.session() as session:
        try:
            result = session.run(cypher_query)
            keys = result.keys()
            records = [record.data() for record in result]
            return keys, records, None
        except Exception as e:
            return [], [], str(e)

def is_valid_password(password):
    return (
        re.search(r'[A-Z]', password) and
        re.search(r'[a-z]', password) and
        re.search(r'[!@#$%^&*()_\-.,?":{}|<>]', password)
    )

def is_valid_email(email):
    email_regex = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return re.match(email_regex, email)

def get_universities_from_csv():
    filepath='open-data/universities.csv'
    try:
        with open(filepath, newline='', encoding='utf-8') as csvfile:
            return [row[0] for row in csv.reader(csvfile)]
    except FileNotFoundError:
        return []

def extract_name_entities(text):
    """Return list of unique name entities (PER, ORG, LOC) from text using spaCy."""
    if not text:
        return []
    doc = nlp(text)
    seen = set()
    ents = []
    for ent in doc.ents:
        if ent.label_ in ("PER", "ORG", "LOC"):
            txt = ent.text.strip()
            if txt not in seen:
                seen.add(txt)
                ents.append(txt)
    return ents

def get_field_list():
    """Read 'Property' column from Excel file, or use defaults."""
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        props = [str(row[0]).strip() for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]
        if props:
            return props
    return [
        "Titel", "Künstler*in/Autor*in", "Sichtbare oder genannte Personen",
        "Entstehungsjahr", "Ort der Entstehung / Nutzung", "Gattung / Genre",
        "Technik", "Dimensionen", "Kurzbeschreibung", "Motive / Topoi",
        "Narrative / Diskurse", "Historischer Kontext",
        "Antiziganistische / Stigmatisierende Elemente", "Agency", "Verknüpfung",
        "Narrativwandel bei Medienwechsel", "Rezeptionsweg", "Sammlung / Archiv",
        "Provenienz", "Literatur", "Ausstellungen / Aufführungen / Veröffentlichungen",
        "Objekt- oder Werkteil", "Rechte / Lizenzen", "Digitalisat-Link/Pfad",
        "Metadaten-Status", "Erfasst von", "Erfassungsdatum", "Kommentar / Anmerkung",
        "Versionsgeschichte"
    ]

# -----------------------
# Routes from the FIRST script (kept intact)
# -----------------------

@app.route("/api/nodes")
def nodes():
    data = get_nodes()
    return jsonify(data)

@app.route("/sections/")
def sections():
    return render_template("sections.html")

# ---------------------------------------------------------
# FETCH ALL PAINTINGS FROM NEO4J
# ---------------------------------------------------------
def get_all_paintings():
    """
    Fetch all paintings from Neo4j, print their full data,
    and return a list of dictionaries for template rendering.
    """
    paintings = []
    with driver.session() as session:
        query = """
            MATCH (p:Painting)
            RETURN p
            ORDER BY p.Name
        """
        result = session.run(query)

        print("\n---- ALL PAINTING NODES IN DATABASE ----\n")

        for record in result:
            node = dict(record["p"])

            # Print full node for debugging
            print("🎨 Painting Node:")
            for k, v in node.items():
                print(f"  {k}: {v}")
            print("--------------------------------------\n")

            # Append for use in template
            paintings.append({
                "name": node.get("Name", ""),
                "title": node.get("Titel", node.get("Name", "")),
                "description": node.get("Kurzbeschreibung", ""),
                "image_path": node.get("Digitalisat-Link/Pfad", ""),
                "url": f"/visual_art/{node.get('Name','')}"
            })

    return paintings


# ---------------------------------------------------------
# VISUAL ART ROUTE
# ---------------------------------------------------------
@app.route("/visual_art",methods=["GET", "POST"])
def visual_art():
    paintings = get_all_paintings()
    return render_template("visual_art.html", paintings=paintings)

@app.route("/book", methods=["GET", "POST"])
def book():
    paintings = get_all_paintings()
    return render_template("book.html", paintings=paintings)

@app.route("/article", methods=["GET", "POST"])
def article():
    paintings = get_all_paintings()
    return render_template("article.html", paintings=paintings)

@app.route("/poem", methods=["GET", "POST"])
def poem():
    paintings = get_all_paintings()
    return render_template("poem.html", paintings=paintings)

@app.route("/song", methods=["GET", "POST"])
def song():
    paintings = get_all_paintings()
    return render_template("song.html", paintings=paintings)

@app.route("/legal_text", methods=["GET", "POST"])
def legal_text():
    paintings = get_all_paintings()
    return render_template("legal_text.html", paintings=paintings)

@app.route("/Person", methods=["GET", "POST"])
def Person():
    paintings = get_all_paintings()
    return render_template("Person.html", paintings=paintings)

@app.route("/poster", methods=["GET", "POST"])
def poster():
    paintings = get_all_paintings()
    return render_template("poster.html", paintings=paintings)

@app.route("/portrait", methods=["GET", "POST"])
def portrait():
    paintings = get_all_paintings()
    return render_template("portrait.html", paintings=paintings)

@app.route("/audio")
def audio():
    paintings = get_all_paintings()
    return render_template("audio.html", paintings=paintings)

@app.route("/video")
def video():
    paintings = get_all_paintings()
    return render_template("video.html", paintings=paintings)
# ---------------------------------------------------------
# INDIVIDUAL PAINTING ROUTE
# ---------------------------------------------------------
@app.route("/visual_art/<painting_name>")
def painting_page(painting_name):
    with driver.session() as session:
        query = """
            MATCH (n:Painting {Name: $name})
            RETURN n
        """
        result = session.run(query, name=painting_name)
        record = result.single()

    if not record:
        return f"No data found for '{painting_name}'."

    node = dict(record["n"])

    return render_template(
        "Individual_page_var.html",
        title=node.get("Titel", painting_name),
        artist_info=node.get("Kurzbeschreibung", ""),
        image_path=node.get("Digitalisat-Link/Pfad", ""),  # correct field
        content=node
    )

@app.route("/book/<painting_name>")
def book(painting_name):
    with driver.session() as session:
        query = """
            MATCH (n:Painting {Name: $name})
            RETURN n
        """
        result = session.run(query, name=painting_name)
        record = result.single()

    if not record:
        return f"No data found for '{painting_name}'."

    node = dict(record["n"])

    return render_template(
        "Individual_page_var.html",
        title=node.get("Titel", painting_name),
        artist_info=node.get("Kurzbeschreibung", ""),
        image_path=node.get("Digitalisat-Link/Pfad", ""),  # correct field
        content=node
    )
@app.route("/article/<painting_name>")
def book(painting_name):
    with driver.session() as session:
        query = """
            MATCH (n:Painting {Name: $name})
            RETURN n
        """
        result = session.run(query, name=painting_name)
        record = result.single()

    if not record:
        return f"No data found for '{painting_name}'."

    node = dict(record["n"])

    return render_template(
        "Individual_page_var.html",
        title=node.get("Titel", painting_name),
        artist_info=node.get("Kurzbeschreibung", ""),
        image_path=node.get("Digitalisat-Link/Pfad", ""),  # correct field
        content=node
    )

@app.route("/poem/<painting_name>")
def book(painting_name):
    with driver.session() as session:
        query = """
            MATCH (n:Painting {Name: $name})
            RETURN n
        """
        result = session.run(query, name=painting_name)
        record = result.single()

    if not record:
        return f"No data found for '{painting_name}'."

    node = dict(record["n"])

    return render_template(
        "Individual_page_var.html",
        title=node.get("Titel", painting_name),
        artist_info=node.get("Kurzbeschreibung", ""),
        image_path=node.get("Digitalisat-Link/Pfad", ""),  # correct field
        content=node
    )

@app.route("/song/<painting_name>")
def book(painting_name):
    with driver.session() as session:
        query = """
            MATCH (n:Painting {Name: $name})
            RETURN n
        """
        result = session.run(query, name=painting_name)
        record = result.single()

    if not record:
        return f"No data found for '{painting_name}'."

    node = dict(record["n"])

    return render_template(
        "Individual_page_var.html",
        title=node.get("Titel", painting_name),
        artist_info=node.get("Kurzbeschreibung", ""),
        image_path=node.get("Digitalisat-Link/Pfad", ""),  # correct field
        content=node
    )

@app.route("/legal_text/<painting_name>")
def book(painting_name):
    with driver.session() as session:
        query = """
            MATCH (n:Painting {Name: $name})
            RETURN n
        """
        result = session.run(query, name=painting_name)
        record = result.single()

    if not record:
        return f"No data found for '{painting_name}'."

    node = dict(record["n"])

    return render_template(
        "Individual_page_var.html",
        title=node.get("Titel", painting_name),
        artist_info=node.get("Kurzbeschreibung", ""),
        image_path=node.get("Digitalisat-Link/Pfad", ""),  # correct field
        content=node
    )

@app.route("/Person/<painting_name>")
def book(painting_name):
    with driver.session() as session:
        query = """
            MATCH (n:Painting {Name: $name})
            RETURN n
        """
        result = session.run(query, name=painting_name)
        record = result.single()

    if not record:
        return f"No data found for '{painting_name}'."

    node = dict(record["n"])

    return render_template(
        "Individual_page_var.html",
        title=node.get("Titel", painting_name),
        artist_info=node.get("Kurzbeschreibung", ""),
        image_path=node.get("Digitalisat-Link/Pfad", ""),  # correct field
        content=node
    )

@app.route("/poster/<painting_name>")
def book(painting_name):
    with driver.session() as session:
        query = """
            MATCH (n:Painting {Name: $name})
            RETURN n
        """
        result = session.run(query, name=painting_name)
        record = result.single()

    if not record:
        return f"No data found for '{painting_name}'."

    node = dict(record["n"])

    return render_template(
        "Individual_page_var.html",
        title=node.get("Titel", painting_name),
        artist_info=node.get("Kurzbeschreibung", ""),
        image_path=node.get("Digitalisat-Link/Pfad", ""),  # correct field
        content=node
    )

@app.route("/portrait/<painting_name>")
def book(painting_name):
    with driver.session() as session:
        query = """
            MATCH (n:Painting {Name: $name})
            RETURN n
        """
        result = session.run(query, name=painting_name)
        record = result.single()

    if not record:
        return f"No data found for '{painting_name}'."

    node = dict(record["n"])

    return render_template(
        "Individual_page_var.html",
        title=node.get("Titel", painting_name),
        artist_info=node.get("Kurzbeschreibung", ""),
        image_path=node.get("Digitalisat-Link/Pfad", ""),  # correct field
        content=node
    )

@app.route("/portrait/<painting_name>")
def book(painting_name):
    with driver.session() as session:
        query = """
            MATCH (n:Painting {Name: $name})
            RETURN n
        """
        result = session.run(query, name=painting_name)
        record = result.single()

    if not record:
        return f"No data found for '{painting_name}'."

    node = dict(record["n"])

    return render_template(
        "Individual_page_var.html",
        title=node.get("Titel", painting_name),
        artist_info=node.get("Kurzbeschreibung", ""),
        image_path=node.get("Digitalisat-Link/Pfad", ""),  # correct field
        content=node
    )

@app.route("/about/")
def about():
    return render_template("about.html")

@app.route("/context-project/")
def context_project():
    return render_template("context_project.html")

@app.route("/history-of-romarchive/")
def history_of_romarchive():
    return render_template("history_of_romarchive.html")

@app.route("/curators/")
def curators():
    return render_template("curators.html")

@app.route("/ethical-guidelines/")
def ethical_guidelines():
    return render_template("ethical_guidelines.html")

@app.route("/collection-policy/")
def collection_policy():
    return render_template("collection_policy.html")

@app.route("/faq/")
def faq():
    return render_template("faq.html")

@app.route("/search/")
def search():
    return render_template("search.html")

@app.route("/terms/")
def terms():
    return render_template("terms.html")

@app.route("/contact/")
def contact():
    return render_template("contact.html")

@app.route("/collection/politics-of-photography/")
def politics_of_photography():
    return render_template("politics_of_photography.html")

@app.route("/visual_art/Individual_page")
def Individual_page():
    return render_template("Individual_page.html")

@app.route("/visual_art/Individual_page_var")
def Individual_page_var():
    section = request.args.get("section", default=None)

    title = ""
    artist_info = """
                   Das Bild ist ambivalent: Es zeigt einerseits Respekt für die Ästhetik und „Malerhaftigkeit“ 
                   der dargestellten Menschen, andererseits reproduziert es stereotype und exotisierende Merkmale. 
                   Es kann sowohl als bewundernde Darstellung als auch als visuelle Festschreibung von „Andersartigkeit“ gelesen werden.
                """
    image_path = "private/Zwei_Zigeuner.png"

    content = ""
    if section == "Objekt_Informationen":
        return render_template("Individual_page_var.html")
    elif section == "Inhaltliche_Beschreibung":
        try:
            with open("s1-templates/zwei_zigeuner_inh_besc.html", "r", encoding="utf-8") as file:
                table_html = file.read()
            content = Markup(table_html)
        except FileNotFoundError:
            content = "Table file not found."
    elif section == "Semantische_Annotation":
        try:
            with open("s1-templates/zwei_zigeuner_sem_ann.html", "r", encoding="utf-8") as file:
                table_html = file.read()
            content = Markup(table_html)
        except FileNotFoundError:
            content = "Table file not found."
    elif section == "Semantische_Relationen":
        try:
            with open("s1-templates/zwei_zigeuner_sem_rel.html", "r", encoding="utf-8") as file:
                table_html = file.read()
            content = Markup(table_html)
        except FileNotFoundError:
            content = "Table file not found."
    elif section == "Technische_rechtliche":
        try:
            with open("s1-templates/zwei_zigeuner_tech_rech.html", "r", encoding="utf-8") as file:
                table_html = file.read()
            content = Markup(table_html)
        except FileNotFoundError:
            content = "Table file not found."
    else:
        return render_template("Individual_page_var.html")

    return render_template(
        "Individual_page_var.html",
        title=title,
        artist_info=artist_info,
        image_path=image_path,
        section=section,
        content=content,
    )

def fetch_object_by_name(name):
    with driver.session() as session:
        result = session.run("MATCH (o:Objekt {name: $name}) RETURN o", name=name)
        record = result.single()
        if record:
            return dict(record["o"])
        return None


@app.route("/index/Metadata/")
def Metadata():
    title = "Metadata Description"
    properties_1 = []
    try:
        with open('data/newdata/description_1.csv', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                properties_1.append((row.get("Property"), row.get("text")))
    except FileNotFoundError:
        properties_1 = []

    properties_2 = []
    try:
        with open('data/newdata/description_2.csv', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                properties_2.append((row.get("Property"), row.get("text")))
    except FileNotFoundError:
        properties_2 = []

    properties_3 = []
    try:
        with open('data/newdata/description_3.csv', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                properties_3.append((row.get("Property"), row.get("text")))
    except FileNotFoundError:
        properties_3 = []

    properties_4 = []
    try:
        with open('data/newdata/description_4.csv', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                properties_4.append((row.get("Property"), row.get("text")))
    except FileNotFoundError:
        properties_4 = []

    properties_5 = []
    try:
        with open('data/newdata/description_5.csv', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                properties_5.append((row.get("Property"), row.get("text")))
    except FileNotFoundError:
        properties_5 = []

    return render_template(
        "Metadata_page_var.html",
        title=title,
        properties_1=properties_1, properties_2=properties_2,
        properties_3=properties_3, properties_4=properties_4,
        properties_5=properties_5
    )

@app.route("/supporters/")
def supporters():
    return "<h1>Supporters Page</h1>"

@app.route("/imprint/")
def imprint():
    return "<h1>Imprint Page</h1>"

@app.route("/privacy/")
def privacy():
    return "<h1>Privacy Statement</h1>"

@app.route("/project-detail")
def project_detail():
    return "Welcome to the Project Detail page!"

@app.route('/delete_all_data')
def delete_all_data():
    with driver.session() as session:
        session.run("MATCH (n) DETACH DELETE n")
    return "All data deleted from Neo4j."

# -----------------------
# Auth / register / login / logout
# -----------------------
@app.route('/register', methods=['GET', 'POST'])
def register():
    universities = get_universities_from_csv()
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        email = request.form.get('email', '')

        if not is_valid_email(email):
            flash('Please enter a valid email address.')
            return redirect(url_for('register'))

        if not is_valid_password(password):
            flash('Password must include at least one uppercase letter, one lowercase letter, and one symbol !@#$%^&*()_-.,?":{}|<>')
            return redirect(url_for('register'))

        name = request.form.get('name')
        institution = request.form.get('institution')
        user_group = request.form.get('user_group')
        password_hash = generate_password_hash(password)

        with driver.session() as session:
            existing_user = session.run("MATCH (u:User {username: $username}) RETURN u", {"username": username}).single()
            existing_email = session.run("MATCH (u:User {email: $email}) RETURN u", {"email": email}).single()

            if existing_user:
                flash("Username already taken.")
                return redirect(url_for('register'))
            if existing_email:
                flash("Email already registered.")
                return redirect(url_for('register'))

            session.run(
                """
                CREATE (u:User {
                    id: toInteger(timestamp()),
                    username: $username,
                    email: $email,
                    password_hash: $password_hash,
                    name: $name,
                    institution: $institution,
                    user_group: $user_group
                })
                """,
                {
                    "username": username,
                    "email": email,
                    "password_hash": password_hash,
                    "name": name,
                    "institution": institution,
                    "user_group": user_group
                }
            )

            flash("Registration successful! You can now log in.")
            return redirect(url_for('login'))

    return render_template('register.html', universities=universities)

# Root/login route(s)
@app.route('/', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')

        with driver.session() as session:
            result = session.run(
                "MATCH (u:User {username: $username}) RETURN u",
                {"username": username}
            )
            record = result.single()
            if record:
                u = record["u"]
                if check_password_hash(u["password_hash"], password):
                    user = User(u.get("id"), u.get("username"), u.get("password_hash"), u.get("name"), u.get("institution"), u.get("user_group"))
                    login_user(user)
                    return redirect(url_for('index'))

        flash("Invalid username or password")
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route("/index")
def index():
    return render_template("index.html")

# -----------------------
# add_data routes (SECOND script integrated under /add_data)
# -----------------------

@app.route('/add_data', methods=['GET'])
@login_required
def add_data():
    fields = get_field_list()
    data, colors, name_entities = {}, {}, {}
    return render_template("objekt_form_buttons.html", fields=fields, data=data, colors=colors, name_entities=name_entities)

@app.route('/add_data/upload', methods=['POST'])
@login_required
def add_data_upload():
    file = request.files.get('file')
    if not file or file.filename == '':
        return "❌ No file selected", 400
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
    file.save(filepath)
    wb = load_workbook(filepath)
    ws = wb.active

    data, colors, name_entities = {}, {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        key = str(row[0]).strip()
        val = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        data[key] = val
        colors[key] = "yellow" if val else "red"
        name_entities[key] = extract_name_entities(val)

    return render_template("objekt_form_buttons.html", fields=list(data.keys()), data=data, colors=colors, name_entities=name_entities)

@app.route('/add_data/extract_entities', methods=['POST'])
@login_required
def add_data_extract_entities():
    fields = get_field_list()
    data, colors, name_entities = {}, {}, {}
    for key in fields:
        val = request.form.get(key, "").strip()
        ne = extract_name_entities(val)
        data[key], name_entities[key] = val, ne
        colors[key] = "green" if ne else "yellow" if val else "red"
    return render_template("objekt_form_buttons.html", fields=fields, data=data, colors=colors, name_entities=name_entities)

@app.route('/add_data/submit', methods=['POST'])
@login_required
def add_data_submit():
    posted = request.form
    rows = []

    # build rows from the posted form
    for key, value in posted.items():
        if key.startswith("custom_property_"):
            index = key.split("_")[-1]
            prop_name = value.strip()
            prop_val = posted.get(f"custom_value_{index}", "").strip() or "No data"
            prop_status = posted.get(f"traffic_custom_{index}", "red")
            if prop_name or prop_val:
                rows.append([prop_name or "No data", prop_val or "No data", prop_status, ""])
        elif key.startswith("traffic_") or key in ("source_node_id", "relation_name", "category", "nodeType"):
            continue  # skip control fields
        else:
            val = value.strip() or "No data"
            status = posted.get(f"traffic_{key}", "red")
            rows.append([key, val, status, ""])

    # final nodeType row
    node_type_value = posted.get("nodeType", posted.get("category", "article"))
    rows.append(["nodeType", node_type_value, "", ""])

    # Build CSV-like string (instead of writing real CSV)
    header = ["Property", "text", "Status", "Name Entity"]
    result_string = "=".join(header) + "\n"

    for row in rows:
        result_string += "=".join(str(col) for col in row) + "\n"
    # Generate unique object ID
    number = random.randint(1, 100)
    category = posted.get("category", "article")
    objekt_id = f"{category}_{number}"

    CSV_DIR = "/app/marburg-project/dataset/german/input/"
    JAR_PATH = "/app/marburg-project/target/QuestionGrammarGenerator.jar"

    source_node_id = posted.get("source_node_id")
    relation_name = posted.get("relation_name")
    # menu = "RELATION" if (source_node_id and relation_name) else "CREATE"

    menu ="CREATE_FROM_STRING"
    neo4j_uri = os.environ.get("NEO4J_URI", "bolt://neo4j:7687")
    neo4j_user = os.environ.get("NEO4J_USER", "neo4j")
    neo4j_pass = os.environ.get("NEO4J_PASSWORD", "password")
    javaflag="FALSE"

    cmd = [
        "java", "-jar", JAR_PATH,
        menu,
        CSV_DIR,  # directory containing generated text file
        neo4j_uri,
        neo4j_user,
        neo4j_pass,
        result_string,
        javaflag
    ]

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, check=True)
        print("stdout:", result.stdout)
        print("stderr:", result.stderr)
    except subprocess.CalledProcessError as e:
        print(e.stderr)
        return f"<h3>ERROR: {e.stderr}</h3>"

    return redirect(url_for('add_data'))

@app.route('/add_data/open_entity_form')
@login_required
def add_data_open_entity_form():
    relation = request.args.get("relation")
    source_node_id = request.args.get("source_node_id")
    entity_name = request.args.get("entity_name")

    fields = get_field_list()
    data, colors, name_entities = {}, {}, {}

    special_field = f"Verknüpfte Entität ({relation})"
    data[special_field] = entity_name
    colors[special_field] = "green"
    name_entities[special_field] = []

    fields = [special_field] + fields
    return render_template("objekt_form_buttons.html", fields=fields, data=data, colors=colors, name_entities=name_entities, source_node_id=source_node_id, relation_name=relation)

# -----------------------
# Optional: ensure default admin user
# -----------------------
def ensure_default_user():
    with driver.session() as session:
        result = session.run("MATCH (u:User {username: $username}) RETURN u", {"username": "admin"})
        if not result.single():
            hashed = generate_password_hash("secret")
            session.run(
                "CREATE (u:User {id: $id, username: $username, password_hash: $password_hash})",
                {"id": 1, "username": "admin", "password_hash": hashed}
            )

# -----------------------
# Run app
# -----------------------
if __name__ == "__main__":
    ensure_default_user()
    app.run(host='0.0.0.0', port=5000, debug=True)
